# -*- coding: utf-8 -*-
"""TSP-4.ipynb

## Traveling salesperson problem
Given a list of cities and the distances between each pair of cities, what is the shortest possible route that visits each city exactly once and returns to the origin city?

**Step 1: Set up and import Gurobi**
"""

# Install Gurobi
!pip install gurobipy
from typing import List

# Import the Gurobi Package
import gurobipy as gp
from gurobipy import GRB

"""**Step 2: Read Excel Data**"""

from openpyxl import load_workbook

wb = load_workbook('TSPdata.xlsx')
ws = wb["Cities"]

cities = [ws.cell(r,1).value for r in range(1,ws.max_row + 1)]
n = len(cities)

xcoord = {ws.cell(r,1).value:ws.cell(r,2).value for r in range(1,ws.max_row + 1)}
ycoord = {ws.cell(r,1).value:ws.cell(r,3).value for r in range(1,ws.max_row + 1)}

# coordinates = {}
# for i in cities:
#   coordinates[i] = (ws.cell(i,2).value, ws.cell(i,3).value)

ws = wb["Distances"]
distance = {}
row = 1
for i in cities:
  for j in cities:
    if i != j:
      distance[i, j] = ws.cell(i,j).value

arcs = distance.keys()

# Display cities
import matplotlib.pyplot as plt
import matplotlib.lines as mlines

for i in cities:
  plt.scatter(xcoord[i], ycoord[i], color='blue')
  plt.text(xcoord[i]+0.1, ycoord[i]+1.1, str(i), fontsize=8)
plt.show()

"""**Step 3: Set up the Linear Programming Model**"""

# Create a Gurobi model
m = gp.Model("TSP")

# Define the decision variables
x = m.addVars(arcs, vtype=GRB.BINARY)
u = m.addVars(range(2, n + 1))

# Define the objective function
obj_expr = sum(distance[i, j]* x[i, j] for i, j in arcs)
m.setObjective(obj_expr, sense=GRB.MINIMIZE)

# Assignment constraints
m.addConstrs(x.sum(i, '*') == 1 for i in cities)
m.addConstrs(x.sum('*', j) == 1 for j in cities)

m.addConstrs(u[i]-u[j]+ n * x[i, j] <= n-1
             for i in range(2, n + 1) for j in range(2, n + 1)
             if i != j
)

m.update()

"""**Step 5: Solve the problem**"""

# Solve the Master Problem
m.optimize()

print(f"Total Distance:{m.ObjVal:6.2f}")

for i in cities:
  plt.scatter(xcoord[i], ycoord[i], color='blue')
  plt.text(xcoord[i]+0.1, ycoord[i]+1.1, str(i), fontsize=8)
  for (i,j) in arcs:
    if x[i,j].X > 0.5:
      plt.plot([xcoord[i], xcoord[j]], [ycoord[i], ycoord[j]], color='black')
plt.show()
