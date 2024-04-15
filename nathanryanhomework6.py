# -*- coding: utf-8 -*-
"""NathanRyanHomework6.ipynb

# Production Planning on Parallel Machines

A manufacturing facility has 5 parallel machines that are used to produce 10 products during the next 28 days.

*   The available time (number of hours) of each machine at each time period in the planning horizon is known. These available times account for maintenance activities.
    *   Machines can work for up to twelve hours of regular time. The sum of regular time and overtime in each period should not exceed the available time. For instance, if the available time for a machine in a particular period is 16 hours, then the machine can be run for twelve hours of regular time and four hours of overtime. The cost of running a machine during regular time and overtime varies per machine but it does not vary per time period.
*   There is a production rate (number of units per hour) associated with each product and machine combination.
    *   The demand for each product is given as number of unit per period.
    *   The on-hand inventory for each product is available at the beginning of the planning horizon and may be used to cover the demand for the first period.
    *   There is no safety stock requirement throughout the planning horizon, however, the company wants the final inventory (i.e., the inventory at the end of the planning horizon) to be at least as large as the beginning inventory.

**Step 1: Set up and import Gurobi**
"""

# Install Gurobi
!pip install gurobipy

# Import the Gurobi Package
import gurobipy as gp
from gurobipy import GRB

"""**Step 2: Read Excel Data**"""

from openpyxl import load_workbook

# Load spreadsheet
wb = load_workbook("Homework6Data.xlsx")

# Read data from the Problem Data worksheet
ws = wb['Problem Data']
nr_machines = ws['B1'].value
nr_products = ws['B2'].value
nr_days = ws['B3'].value
days = range(1,1 + nr_days)
regular_time_limit = ws['B4'].value

# Read data from the Machine Data worksheet
ws = wb['Machine Data']
machines = [ws.cell(2,c).value for c in range(2,2 + nr_machines)]

availability = {(ws.cell(2,c).value, ws.cell(r,1).value) : ws.cell(r,c).value
                for r in range(3,3 + nr_days)
                for c in range(2,2 + nr_machines)
}

regular_cost = {ws.cell(r,8).value : ws.cell(r,9).value
                for r in range(3,3 + nr_machines)
}

overtime_cost = {ws.cell(r,8).value : ws.cell(r,10).value
                 for r in range(3,3 + nr_machines)
}

# Read data from the Product Data worksheet
ws = wb['Product Data']
products = [ws.cell(r,1).value for r in range(3,3 + nr_products)]

production_rate = {(ws.cell(2,c).value, ws.cell(r,1).value) : ws.cell(r,c).value
                   for r in range(3,3 + nr_products)
                   for c in range(2,2 + nr_machines)
}

demand = {ws.cell(r,1).value : ws.cell(r,7).value
          for r in range(3,3 + nr_products)
}

onhand_inventory = {ws.cell(r,1).value : ws.cell(r,8).value
                    for r in range(3,3 + nr_products)
}

availability

"""**Step 3: Set up the Linear Programming Model**"""

#Model
m2 = gp.Model('Parallel Machines 2')

#Vars
x_reg = m2.addVars(machines, days, name = 'x1')
x_over = m2.addVars(machines, days, name = 'x2')
x_prod = m2.addVars(machines, products, days, name = 'x3')
x_inv = m2.addVars(products, days, name = 'x4')

#Obj
reg_hours = sum(x_reg[i,k] * regular_cost[i] for i in machines for k in days )
over_time = sum(x_over[i,k] * overtime_cost[i] for i in machines for k in days)
total = reg_hours + over_time
m2.setObjective(total, sense = GRB.MINIMIZE)

#Constrs
## Time
m2.addConstrs(x_reg[i,k]
              <= regular_time_limit
             for i in machines
             for k in days )

m2.addConstrs(x_reg[i,k] + x_over[i,k]
             <= availability[i,k]
              for i in machines
              for k in days )

#Production
#model.addConstr(production[d, p, m] for j in products == production_rate[m, p] * standard_time[d, m]
m2.addConstrs((x_reg[i,k] + x_over[i,k])
              * production_rate[i,j]
              == x_prod[i,j,k]
              for i in machines
              for j in products
              for k in days)

## Demand/Inventory
for k in range(2,29): #similar method to production planning
    for j in products:
      x_inv[j,k] = x_inv[j,k-1] + sum(x_prod[i,j,k] for i in machines) - demand[j]
      m2.addConstr(x_inv[j,k] >= 0) # this approach allows inventory to be updated over time
                                    # accounting for demand and restricting non-negativity
    #Initial
for j in products:
  x_inv[j,1] = onhand_inventory[j] + sum(x_prod[i,j,1] for i in machines) - demand[j]
  m2.addConstr(x_inv[j,1] >= 0)

    #Final
for j in products:
  m2.addConstr(x_inv[j,28] >= onhand_inventory[j])

'''         Attempt 1                               '''
'''I believe this works.... If I didnt have a variable limit ofc'''
# Create a Gurobi model
m = gp.Model('Parallel Machines')

# Define the decision variables
#machine product combination
x_reg = m.addVars(machines, products, days, name = 'x1') #hours operating of machine reg hours
x_over = m.addVars(machines, products, days, name = 'x2') #hours oper

# Define the objective function
reg_hours = sum(sum(x_reg[i,j,k] for j in products) * regular_cost[i] for i in machines for k in days )
over_time = sum(sum(x_over[i,j,k] for j in products) * overtime_cost[i] for i in machines for k in days)
total = reg_hours + over_time
m.setObjective(total, sense = GRB.MINIMIZE)

# Define the constraints
## Regular and overtime constraints
'''assumption that as we minimize cost and have limited regular time... overtime will not be used until
   all of regular time used as overtime more expensive...'''

m.addConstrs(sum(x_reg[i,j,k] for j in products)
              <= regular_time_limit
             for i in machines
             for k in days )

## Time availability constraints

m.addConstrs(sum(x_reg[i,j,k] for j in products) + sum(x_over[i,j,k] for j in products)
             <= availability[i, k]
              for i in machines
              for k in days )

## Meeting Demands
### Day 1-27
for k in range(1,28):
  m.addConstrs(onhand_inventory[j]
            - x_reg[i,j,k] + x_over[i,j,k]
            * production_rate[i,j]
             == demand[j]
            for i in machines
            for j in products)

## Inventory Balance
### Day 28 ''' this assumse that all inventory has been used... '''
m.addConstrs((x_reg[i,j,28] + x_over[i,j,28])
             * production_rate[i,j]
             == onhand_inventory[j]
             for i in machines
             for j in products)

"""**Step 4: Solve the Model and Display Output**"""

# Solve the model
m2.optimize()

# Display the costs
if m2.status == GRB.OPTIMAL:
    #print(f"Optimal Total Cost: ${m.objVal:.2f}")
    for i in machines:
      for j in products:
        for k in days:
          print(f"Day {k}, machine {i}, product {j}, : {x_prod[i,j,k].x:.2f} units")
else:
    print("NOOOOOOOOOOOOOOOOOO")

m2.write("mymodel.lp")

with open("mymodel.lp", "r") as model_file:
    model_contents = model_file.read()

# Print the model contents
print(model_contents)
